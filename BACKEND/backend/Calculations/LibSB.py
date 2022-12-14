import ast
import random
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from soupsieve.util import lower
from Calculations.models import FileData


def transliterate(text):
    result = ''
    dic = {'Ь': '', 'ь': '', 'Ъ': '', 'ъ': '', 'А': 'A', 'а': 'a', 'Б': 'B', 'б': 'b', 'В': 'V', 'в': 'v',
           'Г': 'G', 'г': 'g', 'Д': 'D', 'д': 'd', 'Е': 'E', 'е': 'e', 'Ё': 'Yo', 'ё': 'yo', 'Ж': 'Zh', 'ж': 'zh',
           'З': 'Z', 'з': 'z', 'И': 'I', 'и': 'i', 'Й': 'Y', 'й': 'y', 'К': 'K', 'к': 'k', 'Л': 'L', 'л': 'l',
           'М': 'M', 'м': 'm', 'Н': 'N', 'н': 'n', 'О': 'O', 'о': 'o', 'П': 'P', 'п': 'p', 'Р': 'R', 'р': 'r',
           'С': 'S', 'с': 's', 'Т': 'T', 'т': 't', 'У': 'U', 'у': 'u', 'Ф': 'F', 'ф': 'f', 'Х': 'H', 'х': 'h',
           'Ц': 'Ts', 'ц': 'ts', 'Ч': 'Ch', 'ч': 'ch', 'Ш': 'Sh', 'ш': 'sh', 'Щ': 'Sch', 'щ': 'sch', 'Ы': 'Yi',
           'ы': 'yi', 'Э': 'E', 'э': 'e', 'Ю': 'Yu', 'ю': 'yu', 'Я': 'Ya', 'я': 'ya', ' ': '_', '8': '8', '9': '9',
           '0': '0', '1': '1', '2': '2', '3': '3', '4': '4', '5': '5', '6': '6', '7': '7', '.': '.','-': '-'}
    for char in text:
        result += dic[char]
    result = result.replace("ul", "ulica").replace('ulitsa', 'ulica').replace('b-r', 'bulvar').replace('bulicavar', 'bulvar').replace('nab', 'naberejnaya')
    if 'pereulok_' in result:
        result = result.replace('pereulok_', '') + '_pereulok'
    return lower(result.replace('.', ''))
def generateID():
    id=''
    for x in range(10):
        id = id + random.choice(
            list('123456789qwertyuiopasdfghjklzxcvbnmQWERTYUIOPASDFGHJKLZXCVBNM_|'))
    return id

def createFile(id_session):
    dataFIle = FileData.objects.get(id_session = id_session)

    book = openpyxl.Workbook()
    sheet = book.active
    sheet["A1"] = "Местоположение"
    sheet["B1"] = "Количество комнат"
    sheet["C1"] = "Сегмент"
    sheet["D1"] = "Этажность дома"
    sheet["E1"] = "Материал стен"
    sheet["F1"] = "Этаж расположения"
    sheet["G1"] = "Площадь квартиры, кв.м"
    sheet["H1"] = "Площадь кухни, кв.м"
    sheet["I1"] = "Наличие балкона/лоджии"
    sheet["J1"] = "Удаленность от станции метро, мин. пешком"
    sheet["K1"] = "Состояние"
    sheet["L1"] = "Цена"

    r = 2
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 30
    sheet.column_dimensions['H'].width = 30
    sheet.column_dimensions['I'].width = 30
    sheet.column_dimensions['J'].width = 30
    sheet.column_dimensions['K'].width = 30
    sheet.column_dimensions['L'].width = 30

    sheet['A1'].fill = PatternFill('solid', fgColor="bdc3c7")
    sheet['B1'].fill = PatternFill('solid', fgColor="bdc3c7")
    sheet['C1'].fill = PatternFill('solid', fgColor="bdc3c7")
    sheet['D1'].fill = PatternFill('solid', fgColor="bdc3c7")
    sheet['E1'].fill = PatternFill('solid', fgColor="bdc3c7")
    sheet['F1'].fill = PatternFill('solid', fgColor="bdc3c7")
    sheet['G1'].fill = PatternFill('solid', fgColor="bdc3c7")
    sheet['H1'].fill = PatternFill('solid', fgColor="bdc3c7")
    sheet['I1'].fill = PatternFill('solid', fgColor="bdc3c7")
    sheet['J1'].fill = PatternFill('solid', fgColor="bdc3c7")
    sheet['K1'].fill = PatternFill('solid', fgColor="bdc3c7")
    sheet['L1'].fill = PatternFill('solid', fgColor="bdc3c7")

    for cell in ast.literal_eval(dataFIle.NewData):
        sheet[r][0].value =cell["location"]
        sheet[r][1].value =cell["numRooms"]
        sheet[r][2].value =cell["segment"]
        sheet[r][3].value =cell["floorsHouse"]
        sheet[r][4].value =cell["materialWall"]
        sheet[r][5].value =cell["floor"]
        sheet[r][6].value =cell["areaApart"]
        sheet[r][7].value =cell["areaKitchen"]
        sheet[r][8].value =cell["balcony"]
        sheet[r][9].value =cell["proxMetro"]
        sheet[r][10].value =cell["structure"]
        sheet[r][11].value =cell["price"]
        r+=1
    name = f"static/Export_{id_session.replace('|', '')}.xlsx"
    print(name, id_session)
    book.save(name)

    book.close()
    s = FileData.objects.get(id_session = id_session)
    s.EndPool = name
    s.save()
    return 0
