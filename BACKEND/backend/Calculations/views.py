import ast
import json
import xlrd
from rest_framework.response import Response
from rest_framework.views import APIView
from Calculations.models import FileData
import random
import openpyxl

from Calculations.price_calculation import calculate_price
from Calculations.selectionAnalogues import AnalogsMirCvartir, AnalogsMove, sortingAnalogs


class LoadFile(APIView):
    def post(self, request):
        id_session = ''
        for x in range(32):
            id_session = id_session + random.choice(
                list('123456789qwertyuiopasdfghjklzxcvbnmQWERTYUIOPASDFGHJKLZXCVBNM_|'))
        pr = FileData(
            file=request.data.get('file'),
            id_user=request.data.get('id_user'),
            path='',
            id_session=id_session,
        )
        pr.save()
        obj = FileData.objects.get(id_session=id_session)
        obj.path = pr.file.name
        obj.save()
        return Response({"id_session": id_session, "data": read_open(id_session)})


def read_open(id_session):
    lot = []
    id = 0
    path = FileData.objects.get(id_session=id_session).path
    try:
        wookbook = openpyxl.load_workbook(path)
        worksheet = wookbook.active
        start_position_y = 0
        for x in range(1, worksheet.max_row + 1):
            if ((worksheet.cell(row=x, column=1).value != None) &
                    (worksheet.cell(row=x, column=1).value == 'Местоположение')):
                start_position_y = x + 1
        for y in range(start_position_y, worksheet.max_row + 1):
            if (worksheet.cell(row=y, column=1).value != None):
                data = {
                    "id_Apart": id,
                    "location": worksheet.cell(row=y, column=1).value+'.',
                    "numRooms": worksheet.cell(row=y, column=2).value,
                    "segment": worksheet.cell(row=y, column=3).value,
                    "floorsHouse": worksheet.cell(row=y, column=4).value,
                    "materialWall": worksheet.cell(row=y, column=5).value,
                    "floor": worksheet.cell(row=y, column=6).value,
                    "areaApart": worksheet.cell(row=y, column=7).value,
                    "areaKitchen": worksheet.cell(row=y, column=8).value,
                    "balcony": worksheet.cell(row=y, column=9).value,
                    "proxMetro": worksheet.cell(row=y, column=10).value,
                    "structure": worksheet.cell(row=y, column=11).value,
                    "coordinates": '',
                }
                id += 1
                lot.append(data)
    except:
        wookbook = xlrd.open_workbook(path ,formatting_info=True)
        worksheet = wookbook.sheet_by_index(0)
        start_position_y = 0
        for x in range(0, worksheet.nrows):
            if ((worksheet.cell(rowx=x, colx=0).value != None) & (worksheet.cell(rowx=x, colx=0).value == 'Местоположение')):
                start_position_y = x + 1
        for y in range(start_position_y, worksheet.nrows):
            if (worksheet.cell(rowx=y, colx=0).value != None):
                data = {
                    "id_Apart": id,
                    "location": worksheet.cell(rowx=y, colx=0).value+'.',
                    "numRooms": worksheet.cell(rowx=y, colx=1).value,
                    "segment": worksheet.cell(rowx=y, colx=2).value,
                    "floorsHouse": worksheet.cell(rowx=y, colx=3).value,
                    "materialWall": worksheet.cell(rowx=y, colx=4).value,
                    "floor": worksheet.cell(rowx=y, colx=5).value,
                    "areaApart": worksheet.cell(rowx=y, colx=6).value,
                    "areaKitchen": worksheet.cell(rowx=y, colx=7).value,
                    "balcony": worksheet.cell(rowx=y, colx=8).value,
                    "proxMetro": worksheet.cell(rowx=y, colx=9).value,
                    "structure": worksheet.cell(rowx=y, colx=10).value,
                    "coordinates": '',
                }
                id += 1
                lot.append(data)
    obj = FileData.objects.get(id_session=id_session)
    obj.data = lot
    obj.save()
    return lot


class selectionAnalogs(APIView):

    def post(self, request, format=None):
        id_session = request.data.get('id_session')
        id_referens = request.data.get('id_ref')
        print(id_session, id_referens)
        arrFile = FileData.objects.get(id_session=id_session)
        arrFile.id_reference = id_referens
        arrFile.save()
        return Response("0")

    def get(self, request, format=None):
        Apart = []
        GlobalApart = []
        id_session = self.request.query_params.get('id_session')
        DataSession = FileData.objects.filter(id_session=id_session)[0]
        data = ast.literal_eval(DataSession.id_reference)
        print(data)
        print("start parsing")
        for id_reference in data:
            Apart.append({"analog":AnalogsMirCvartir(id_session, int(id_reference))})
            Apart.append({"analog":AnalogsMove(id_session, int(id_reference))})
        print("end parsing")
        for analog in Apart:
            for info in analog["analog"]:
                GlobalApart.append(info)
        print("start sort")
        print(GlobalApart)
        list = []
        for id_reference in data:
            list.append({"resualt": {"data": (sortingAnalogs(GlobalApart, id_session, id_reference)), "id": id_reference}})
        saveAnalogs = FileData.objects.get(id_session=id_session)
        saveAnalogs.data_analogs = list
        saveAnalogs.save()
        return Response(json.dumps(list))
#
# class AllInfoSession(APIView):
#     def get(self, request, format=None):
#         id_session = self.request.query_params.get('id_session')
#         info = FileData.objects.get(id_session=id_session)
#         data = {
#             "id_session": info.id_session,
#             "data": json.dumps(info.data),
#             "id_reference": info.id_reference,
#             "data_analogs": json.dumps(info.data_analogs),
#             "ids_analogs": json.dumps(info.ids_analogs),
#         }
#         return Response(data)

class Calculation(APIView):
    def post(self, request, format=False):
        reports = []
        id_session = self.request.data.get('id_session')
        ids_analogs = []
        ids = self.request.data.get('ids_analogs').split(',')
        for id in ids:
            ids_analogs.append(id)
        data = FileData.objects.get(id_session=id_session)
        data.ids_analogs = ids_analogs
        data.save()
        reports.append(calculate_price(id_session))
        data = FileData.objects.get(id_session=id_session)
        data.report = reports
        data.save()
        print(reports)
        return Response(reports)
